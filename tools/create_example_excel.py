"""
Create example Excel file with proper format for the combination engine.
This script generates an Excel file with the correct column structure.
"""
import openpyxl
from pathlib import Path

# Create a new workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Test Cases"

# Define the column headers and sample data
headers = [
    # API Configuration
    "[API]endpoint",
    "[API]Method",

    # Request Headers
    "[Request][Header]Content-Type",
    "[Request][Header]Authorization",
    "[Request][Header]X-Mock-Status",

    # Request Body - Top level
    "[Request][Body]referenceId",
    "[Request][Body]sourceSystem",
    "[Request][Body]channel",

    # Request Body - Agent
    "[Request][Body]data.agent.agentCode",
    "[Request][Body]data.agent.agentName",
    "[Request][Body]data.agent.businessSource",
    "[Request][Body]data.agent.channel",
    "[Request][Body]data.agent.masLicenseNo",
    "[Request][Body]data.agent.trainingCodes[0]",

    # Request Body - Client Profiles[0] (ML)
    "[Request][Body]data.clientProfiles[0].id",
    "[Request][Body]data.clientProfiles[0].clientType",
    "[Request][Body]data.clientProfiles[0].relationship",
    "[Request][Body]data.clientProfiles[0].personalInfo.isSmoker[Type:bool]",
    "[Request][Body]data.clientProfiles[0].personalInfo.name",
    "[Request][Body]data.clientProfiles[0].personalInfo.dob",
    "[Request][Body]data.clientProfiles[0].personalInfo.age[Type:int]",
    "[Request][Body]data.clientProfiles[0].personalInfo.gender",
    "[Request][Body]data.clientProfiles[0].personalInfo.isImpairedLife[Type:bool]",
    "[Request][Body]data.clientProfiles[0].personalInfo.premiumBackDate",
    "[Request][Body]data.clientProfiles[0].personalInfo.residentialStatus",
    "[Request][Body]data.clientProfiles[0].personalInfo.nationality.code",
    "[Request][Body]data.clientProfiles[0].personalInfo.nationality.value",
    "[Request][Body]data.clientProfiles[0].personalInfo.countryOfBirth.code",
    "[Request][Body]data.clientProfiles[0].personalInfo.countryOfBirth.value",
    "[Request][Body]data.clientProfiles[0].personalInfo.countryOfResidence.code",
    "[Request][Body]data.clientProfiles[0].personalInfo.countryOfResidence.value",
    "[Request][Body]data.clientProfiles[0].personalInfo.passType.code",
    "[Request][Body]data.clientProfiles[0].personalInfo.passType.value",
    "[Request][Body]data.clientProfiles[0].personalInfo.occupation.code",
    "[Request][Body]data.clientProfiles[0].personalInfo.occupation.value",
    "[Request][Body]data.clientProfiles[0].personalInfo.occupation.class",

    # Request Body - Client Profiles[1] (O)
    "[Request][Body]data.clientProfiles[1].id",
    "[Request][Body]data.clientProfiles[1].clientType",
    "[Request][Body]data.clientProfiles[1].relationship",
    "[Request][Body]data.clientProfiles[1].personalInfo.isSmoker[Type:bool]",
    "[Request][Body]data.clientProfiles[1].personalInfo.name",
    "[Request][Body]data.clientProfiles[1].personalInfo.dob",
    "[Request][Body]data.clientProfiles[1].personalInfo.age[Type:int]",
    "[Request][Body]data.clientProfiles[1].personalInfo.gender",
    "[Request][Body]data.clientProfiles[1].personalInfo.isImpairedLife[Type:bool]",
    "[Request][Body]data.clientProfiles[1].personalInfo.premiumBackDate",
    "[Request][Body]data.clientProfiles[1].personalInfo.residentialStatus",
    "[Request][Body]data.clientProfiles[1].personalInfo.nationality.code",
    "[Request][Body]data.clientProfiles[1].personalInfo.nationality.value",
    "[Request][Body]data.clientProfiles[1].personalInfo.countryOfBirth.code",
    "[Request][Body]data.clientProfiles[1].personalInfo.countryOfBirth.value",
    "[Request][Body]data.clientProfiles[1].personalInfo.countryOfResidence.code",
    "[Request][Body]data.clientProfiles[1].personalInfo.countryOfResidence.value",
    "[Request][Body]data.clientProfiles[1].personalInfo.passType.code",
    "[Request][Body]data.clientProfiles[1].personalInfo.passType.value",
    "[Request][Body]data.clientProfiles[1].personalInfo.occupation.code",
    "[Request][Body]data.clientProfiles[1].personalInfo.occupation.value",
    "[Request][Body]data.clientProfiles[1].personalInfo.occupation.class",

    # Request Body - Filters
    "[Request][Body]data.filters.isCrossLife[Type:bool]",
    "[Request][Body]data.filters.isJointLife[Type:bool]",
    "[Request][Body]data.filters.isJointProposer[Type:bool]",
    "[Request][Body]data.filters.isPayerSecurity[Type:bool]",
    "[Request][Body]data.filters.isCrisisWaiver[Type:bool]",

    # Response expectations
    "[Response][API]status",
    "[Response][Body]success[Type:bool]",
    "[Response][Body]referenceId",
    "[Response][Body]sourceSystem",
]

# Sample data row
data_row = [
    # API Configuration
    "https://mockoon-api.techlabth.com/api/combination-data",
    "POST",

    # Request Headers
    "application/json",
    "Bearer your-token-here",
    "200",

    # Request Body - Top level
    "a7f82adc-75cd-4fa6-90b1-4e98f7e04fcb",
    "PRUJOURNEY",
    "AG",

    # Request Body - Agent
    "21755",
    "QA AD Channel Testing Client Ten",
    "AG",
    "AD",
    "",
    "PMUM",

    # Request Body - Client Profiles[0] (ML)
    "7e3fe367-4acc-4bc1-a522-461ab94ecfc9",
    "ML",
    "OS",
    "false",
    "null null namkeen",
    "15.04.1998",
    "27",
    "F",
    "false",
    "[NULL]",
    "[NULL]",
    "SNG",
    "[NULL]",
    "[NULL]",
    "[NULL]",
    "SNG",
    "[NULL]",
    "[NULL]",
    "[NULL]",
    "ABBT",
    "[NULL]",
    "1",

    # Request Body - Client Profiles[1] (O)
    "7e3fe367-4acc-4bc1-a522-461ab94ecfc9",
    "O",
    "OS",
    "false",
    "null null namkeen",
    "15.04.1998",
    "27",
    "F",
    "false",
    "[NULL]",
    "[NULL]",
    "SNG",
    "[NULL]",
    "[NULL]",
    "[NULL]",
    "SNG",
    "[NULL]",
    "[NULL]",
    "[NULL]",
    "ABBT",
    "[NULL]",
    "1",

    # Request Body - Filters
    "true",
    "false",
    "false",
    "false",
    "false",

    # Response expectations
    "200",
    "true",
    "a7f82adc-75cd-4fa6-90b1-4e98f7e04fcb",
    "PRUJOURNEY",
]

# Write headers
for col_idx, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = openpyxl.styles.Font(bold=True)
    cell.fill = openpyxl.styles.PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

# Write data row
for col_idx, value in enumerate(data_row, start=1):
    ws.cell(row=2, column=col_idx, value=value)

# Auto-adjust column widths
for column in ws.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = min(max_length + 2, 50)  # Cap at 50
    ws.column_dimensions[column_letter].width = adjusted_width

# Save the workbook
output_path = Path(__file__).parent.parent / "data" / "example-test-cases.xlsx"
wb.save(output_path)
print(f"✅ Excel file created: {output_path}")
print(f"   Total columns: {len(headers)}")
print(f"   Sample data rows: 1")
