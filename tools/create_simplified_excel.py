"""
Create simplified Excel file using array expansion notation [].
This demonstrates the new feature where comma-separated values create multiple array items.
"""
import openpyxl
from pathlib import Path

# Create a new workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Test Cases"

# Define the column headers using [] for array expansion
headers = [
    # API Configuration
    "[API]endpoint",
    "[API]Method",

    # Request Headers
    "[Request][Header]Content-Type",
    "[Request][Header]X-Mock-Status",

    # Request Body - Top level
    "[Request][Body]referenceId",
    "[Request][Body]sourceSystem",
    "[Request][Body]channel",

    # Request Body - Agent (single object, no array)
    "[Request][Body]data.agent.agentCode",
    "[Request][Body]data.agent.agentName",
    "[Request][Body]data.agent.businessSource",
    "[Request][Body]data.agent.channel",
    "[Request][Body]data.agent.masLicenseNo",
    "[Request][Body]data.agent.trainingCodes[0]",

    # Request Body - Client Profiles (using [] for expansion)
    # This single field with comma-separated values will create 2 array items
    "[Request][Body]data.clientProfiles[].id",
    "[Request][Body]data.clientProfiles[].clientType",  # "ML,O" → creates 2 items
    "[Request][Body]data.clientProfiles[].relationship",

    # Personal Info (shared across all array items)
    "[Request][Body]data.clientProfiles[].personalInfo.isSmoker[Type:bool]",
    "[Request][Body]data.clientProfiles[].personalInfo.name",
    "[Request][Body]data.clientProfiles[].personalInfo.dob",
    "[Request][Body]data.clientProfiles[].personalInfo.age[Type:int]",
    "[Request][Body]data.clientProfiles[].personalInfo.gender",
    "[Request][Body]data.clientProfiles[].personalInfo.isImpairedLife[Type:bool]",
    "[Request][Body]data.clientProfiles[].personalInfo.premiumBackDate",
    "[Request][Body]data.clientProfiles[].personalInfo.residentialStatus",

    # Nationality
    "[Request][Body]data.clientProfiles[].personalInfo.nationality.code",
    "[Request][Body]data.clientProfiles[].personalInfo.nationality.value",

    # Country of Birth
    "[Request][Body]data.clientProfiles[].personalInfo.countryOfBirth.code",
    "[Request][Body]data.clientProfiles[].personalInfo.countryOfBirth.value",

    # Country of Residence
    "[Request][Body]data.clientProfiles[].personalInfo.countryOfResidence.code",
    "[Request][Body]data.clientProfiles[].personalInfo.countryOfResidence.value",

    # Pass Type
    "[Request][Body]data.clientProfiles[].personalInfo.passType.code",
    "[Request][Body]data.clientProfiles[].personalInfo.passType.value",

    # Occupation
    "[Request][Body]data.clientProfiles[].personalInfo.occupation.code",
    "[Request][Body]data.clientProfiles[].personalInfo.occupation.value",
    "[Request][Body]data.clientProfiles[].personalInfo.occupation.class",

    # Request Body - Filters
    "[Request][Body]data.filters.isCrossLife[Type:bool]",
    "[Request][Body]data.filters.isJointLife[Type:bool]",
    "[Request][Body]data.filters.isJointProposer[Type:bool]",
    "[Request][Body]data.filters.isPayerSecurity[Type:bool]",
    "[Request][Body]data.filters.isCrisisWaiver[Type:bool]",

    # Response expectations
    "[Response][API]status",
    "[Response][Body]success[Type:bool]",
    "[Response][Body]referenceId",
    "[Response][Body]sourceSystem",

    # Response - Array search validation
    "[Response][Body]data.eligibleProducts[].prodCode",
]

# Sample data row - using comma-separated clientType to create 2 array items
data_row = [
    # API Configuration
    "https://mockoon-api.techlabth.com/api/combination-data",
    "POST",

    # Request Headers
    "application/json",
    "200",

    # Request Body - Top level
    "a7f82adc-75cd-4fa6-90b1-4e98f7e04fcb",
    "PRUJOURNEY",
    "AG",

    # Request Body - Agent
    "21755",
    "QA AD Channel Testing Client Ten",
    "AG",
    "AD",
    "",
    "PMUM",

    # Request Body - Client Profiles (using [] expansion)
    # Same ID for both clients
    "7e3fe367-4acc-4bc1-a522-461ab94ecfc9",
    # COMMA-SEPARATED: This creates 2 array items with different clientType
    "ML,O",
    # Same relationship for both
    "OS",

    # Personal Info (same for both clients)
    "false",
    "null null namkeen",
    "15.04.1998",
    "27",
    "F",
    "false",
    "[NULL]",
    "[NULL]",

    # Nationality (same for both)
    "SNG",
    "[NULL]",

    # Country of Birth (same for both)
    "[NULL]",
    "[NULL]",

    # Country of Residence (same for both)
    "SNG",
    "[NULL]",

    # Pass Type (same for both)
    "[NULL]",
    "[NULL]",

    # Occupation (same for both)
    "ABBT",
    "[NULL]",
    "1",

    # Request Body - Filters
    "true",
    "false",
    "false",
    "false",
    "false",

    # Response expectations
    "200",
    "true",
    "a7f82adc-75cd-4fa6-90b1-4e98f7e04fcb",
    "PRUJOURNEY",

    # Response - Array search validation (expect to find this prodCode in the array)
    "CT6",
]

# Write headers
for col_idx, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = openpyxl.styles.Font(bold=True)
    cell.fill = openpyxl.styles.PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

# Write data row
for col_idx, value in enumerate(data_row, start=1):
    ws.cell(row=2, column=col_idx, value=value)

# Highlight the expansion column (clientType with ML,O)
clientType_col = headers.index("[Request][Body]data.clientProfiles[].clientType") + 1
ws.cell(row=2, column=clientType_col).fill = openpyxl.styles.PatternFill(
    start_color="FFFF00", end_color="FFFF00", fill_type="solid"
)
# Add comment
ws.cell(row=2, column=clientType_col).comment = openpyxl.comments.Comment(
    "Comma-separated values create multiple array items!\n\n"
    "ML,O → Creates 2 objects:\n"
    "- clientProfiles[0].clientType = 'ML'\n"
    "- clientProfiles[1].clientType = 'O'\n\n"
    "All other fields are duplicated across both items.",
    "Claude Code"
)

# Auto-adjust column widths
for column in ws.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = min(max_length + 2, 50)  # Cap at 50
    ws.column_dimensions[column_letter].width = adjusted_width

# Save the workbook
output_path = Path(__file__).parent.parent / "data" / "example-simplified.xlsx"
wb.save(output_path)
print(f"✅ Simplified Excel file created: {output_path}")
print(f"   Total columns: {len(headers)} (reduced from 67!)")
print(f"   Sample data rows: 1")
print(f"")
print(f"Key feature:")
print(f"   [Request][Body]data.clientProfiles[].clientType = 'ML,O'")
print(f"   → Creates 2 array items with all fields duplicated")
print(f"   → clientProfiles[0].clientType = 'ML'")
print(f"   → clientProfiles[1].clientType = 'O'")
