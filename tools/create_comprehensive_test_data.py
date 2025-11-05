"""
Create comprehensive Excel test data file with multiple test cases
Based on the provided screenshots showing various test scenarios
"""
import openpyxl
from pathlib import Path

# Create a new workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Test Cases"

# Define the column headers
headers = [
    # API Configuration
    "[API]endpoint",
    "[API]Method",

    # Request Body
    "[Request][Body]referenceId",
    "[Request][Body]sourceSystem",
    "[Request][Body]channel",

    # Agent data
    "[Request][Body]data.agent.agentCode",
    "[Request][Body]data.agent.agentName",
    "[Request][Body]data.agent.channel",
    "[Request][Body]data.agent.masLicenseNo",
    "[Request][Body]data.agent.trainingCodes[0]",

    # Client Profiles - using [] for array expansion
    "[Request][Body]data.clientProfiles[].clientType",
    "[Request][Body]data.clientProfiles[].personalInfo.isSmoker[Type:bool]",
    "[Request][Body]data.clientProfiles[].personalInfo.dob",
    "[Request][Body]data.clientProfiles[].personalInfo.age[Type:int]",
    "[Request][Body]data.clientProfiles[].personalInfo.premiumBackDate[Type:bool]",
    "[Request][Body]data.clientProfiles[].personalInfo.countryOfResidence.code",
    "[Request][Body]data.clientProfiles[].personalInfo.passType.code",
    "[Request][Body]data.clientProfiles[].personalInfo.passType.value",
    "[Request][Body]data.clientProfiles[].personalInfo.occupation.code",
    "[Request][Body]data.clientProfiles[].personalInfo.occupation.value",
    "[Request][Body]data.clientProfiles[].personalInfo.occupation.class",

    # Filters
    "[Request][Body]data.filters.isCrossLife[Type:bool]",
    "[Request][Body]data.filters.isPayerSecurity[Type:bool]",
    "[Request][Body]data.filters.isCrisisWaiver[Type:bool]",

    # Response expectations
    "[Response][Body]sourceSystem",
    "[Response][Body]data.eligibleProducts[].prodCode",
    "[Response][Body]data.nonEligibleProducts[].prodCode",
]

# Test data variations - different filter combinations
test_scenarios = [
    # Scenario 1: Basic test
    {"isCrossLife": "TRUE", "isPayerSecurity": "TRUE", "isCrisisWaiver": "FALSE", "age": "16", "premiumBackDate": "TRUE"},
    {"isCrossLife": "TRUE", "isPayerSecurity": "TRUE", "isCrisisWaiver": "FALSE", "age": "16", "premiumBackDate": "TRUE"},
    {"isCrossLife": "TRUE", "isPayerSecurity": "FALSE", "isCrisisWaiver": "TRUE", "age": "16", "premiumBackDate": "TRUE"},
    {"isCrossLife": "TRUE", "isPayerSecurity": "FALSE", "isCrisisWaiver": "FALSE", "age": "16", "premiumBackDate": "TRUE"},
    {"isCrossLife": "FALSE", "isPayerSecurity": "TRUE", "isCrisisWaiver": "TRUE", "age": "16", "premiumBackDate": "TRUE"},
    {"isCrossLife": "FALSE", "isPayerSecurity": "TRUE", "isCrisisWaiver": "FALSE", "age": "16", "premiumBackDate": "TRUE"},
    {"isCrossLife": "FALSE", "isPayerSecurity": "FALSE", "isCrisisWaiver": "TRUE", "age": "16", "premiumBackDate": "TRUE"},
    {"isCrossLife": "FALSE", "isPayerSecurity": "FALSE", "isCrisisWaiver": "FALSE", "age": "16", "premiumBackDate": "TRUE"},
    {"isCrossLife": "TRUE", "isPayerSecurity": "TRUE", "isCrisisWaiver": "FALSE", "age": "16", "premiumBackDate": "FALSE"},
    {"isCrossLife": "TRUE", "isPayerSecurity": "TRUE", "isCrisisWaiver": "FALSE", "age": "16", "premiumBackDate": "FALSE"},
    {"isCrossLife": "TRUE", "isPayerSecurity": "FALSE", "isCrisisWaiver": "TRUE", "age": "16", "premiumBackDate": "FALSE"},
    {"isCrossLife": "TRUE", "isPayerSecurity": "FALSE", "isCrisisWaiver": "FALSE", "age": "16", "premiumBackDate": "FALSE"},
    {"isCrossLife": "FALSE", "isPayerSecurity": "TRUE", "isCrisisWaiver": "TRUE", "age": "16", "premiumBackDate": "FALSE"},
    {"isCrossLife": "FALSE", "isPayerSecurity": "TRUE", "isCrisisWaiver": "FALSE", "age": "16", "premiumBackDate": "FALSE"},
    {"isCrossLife": "FALSE", "isPayerSecurity": "FALSE", "isCrisisWaiver": "TRUE", "age": "16", "premiumBackDate": "FALSE"},
    {"isCrossLife": "FALSE", "isPayerSecurity": "FALSE", "isCrisisWaiver": "FALSE", "age": "16", "premiumBackDate": "FALSE"},
    # Age 17 scenarios
    {"isCrossLife": "TRUE", "isPayerSecurity": "TRUE", "isCrisisWaiver": "FALSE", "age": "17", "premiumBackDate": "TRUE"},
    {"isCrossLife": "TRUE", "isPayerSecurity": "TRUE", "isCrisisWaiver": "FALSE", "age": "17", "premiumBackDate": "TRUE"},
    {"isCrossLife": "TRUE", "isPayerSecurity": "FALSE", "isCrisisWaiver": "TRUE", "age": "17", "premiumBackDate": "TRUE"},
    {"isCrossLife": "TRUE", "isPayerSecurity": "FALSE", "isCrisisWaiver": "FALSE", "age": "17", "premiumBackDate": "TRUE"},
    {"isCrossLife": "FALSE", "isPayerSecurity": "TRUE", "isCrisisWaiver": "TRUE", "age": "17", "premiumBackDate": "TRUE"},
    {"isCrossLife": "FALSE", "isPayerSecurity": "TRUE", "isCrisisWaiver": "FALSE", "age": "17", "premiumBackDate": "TRUE"},
    {"isCrossLife": "FALSE", "isPayerSecurity": "FALSE", "isCrisisWaiver": "TRUE", "age": "17", "premiumBackDate": "TRUE"},
    {"isCrossLife": "FALSE", "isPayerSecurity": "FALSE", "isCrisisWaiver": "FALSE", "age": "17", "premiumBackDate": "TRUE"},
    {"isCrossLife": "TRUE", "isPayerSecurity": "TRUE", "isCrisisWaiver": "FALSE", "age": "17", "premiumBackDate": "FALSE"},
    {"isCrossLife": "TRUE", "isPayerSecurity": "TRUE", "isCrisisWaiver": "FALSE", "age": "17", "premiumBackDate": "FALSE"},
    {"isCrossLife": "TRUE", "isPayerSecurity": "FALSE", "isCrisisWaiver": "TRUE", "age": "17", "premiumBackDate": "FALSE"},
    {"isCrossLife": "TRUE", "isPayerSecurity": "FALSE", "isCrisisWaiver": "FALSE", "age": "17", "premiumBackDate": "FALSE"},
    {"isCrossLife": "FALSE", "isPayerSecurity": "TRUE", "isCrisisWaiver": "TRUE", "age": "17", "premiumBackDate": "FALSE"},
    {"isCrossLife": "FALSE", "isPayerSecurity": "TRUE", "isCrisisWaiver": "FALSE", "age": "17", "premiumBackDate": "FALSE"},
    {"isCrossLife": "FALSE", "isPayerSecurity": "FALSE", "isCrisisWaiver": "TRUE", "age": "17", "premiumBackDate": "FALSE"},
    {"isCrossLife": "FALSE", "isPayerSecurity": "FALSE", "isCrisisWaiver": "FALSE", "age": "17", "premiumBackDate": "FALSE"},
    # Age 50
    {"isCrossLife": "TRUE", "isPayerSecurity": "TRUE", "isCrisisWaiver": "TRUE", "age": "50", "premiumBackDate": "TRUE"},
]

# Write headers
for col_idx, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = openpyxl.styles.Font(bold=True)
    cell.fill = openpyxl.styles.PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

# Write test data rows
for row_idx, scenario in enumerate(test_scenarios, start=2):
    data_row = [
        # API Configuration
        "https://vpms-dev.aks.lb1-sglife-uat-az1-ios0n9.pru.intranet.asia/vpms-api/v2/products",
        "POST",

        # Request Body - Top level
        "",  # referenceId - blank
        "PRUJOURNEY",
        "AG",

        # Agent data
        "21755",
        "QA AD Channel Testing Client Ten",
        "AD",
        "[EMPTY]",
        "PVLI",

        # Client Profiles
        "ML",
        "TRUE",
        "02.09.1998",
        scenario["age"],
        scenario["premiumBackDate"],
        "SNG",
        "[NULL]",
        "[NULL]",
        "[NULL]",
        "[NULL]",
        "[NULL]",

        # Filters
        scenario["isCrossLife"],
        scenario["isPayerSecurity"],
        scenario["isCrisisWaiver"],

        # Response expectations
        "PRUJOURNEY",
        "",  # eligibleProducts - leave blank for some rows
        "PVLI",  # nonEligibleProducts - expect PVLI
    ]

    for col_idx, value in enumerate(data_row, start=1):
        ws.cell(row=row_idx, column=col_idx, value=value)

# Auto-adjust column widths
for column in ws.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        try:
            if cell.value and len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = min(max_length + 2, 60)  # Cap at 60
    ws.column_dimensions[column_letter].width = adjusted_width

# Save the workbook
output_path = Path(__file__).parent.parent / "data" / "comprehensive-test-data.xlsx"
wb.save(output_path)
print(f"✅ Comprehensive test data file created: {output_path}")
print(f"   Total test cases: {len(test_scenarios)}")
print(f"   Total columns: {len(headers)}")
print(f"")
print(f"Key features:")
print(f"   - Multiple filter combinations (isCrossLife, isPayerSecurity, isCrisisWaiver)")
print(f"   - Different age scenarios (16, 17, 50)")
print(f"   - Premium back date variations (TRUE/FALSE)")
print(f"   - Array validation: nonEligibleProducts[].prodCode should contain 'PVLI'")
